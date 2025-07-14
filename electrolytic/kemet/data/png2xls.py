# coding=utf-8


def_baidu=False
def_tencent=True

import numpy as np
import openpyxl
import pandas as pd
if def_baidu:
    import sys
    import json
    import base64

    # 保证兼容python2以及python3
    IS_PY3 = sys.version_info.major == 3
    if IS_PY3:
        from urllib.request import urlopen
        from urllib.request import Request
        from urllib.error import URLError
        from urllib.parse import urlencode
        from urllib.parse import quote_plus
    else:
        import urllib2
        from urllib import quote_plus
        # from urllib2 import urlopen
        # from urllib2 import Request
        # from urllib2 import URLError
        from urllib import urlencode

    # 防止https证书校验不正确
    import ssl
    ssl._create_default_https_context = ssl._create_unverified_context

if def_tencent:
    import os
    import json
    import base64
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
    from tencentcloud.ocr.v20181119 import ocr_client, models



# API_KEY = 'F8Go5rLbjMozTcWE5yClhKCa'
#
# SECRET_KEY = 'm9Ey40uL6mE4N45Hndr8IdqANImbaVjB'
#
#
# #OCR_URL = "https://aip.baidubce.com/rest/2.0/ocr/v1/accurate_basic"
# OCR_URL ="https://aip.baidubce.com/rest/2.0/ocr/v1/table"
#
# """  TOKEN start """
# TOKEN_URL = 'https://aip.baidubce.com/oauth/2.0/token'


"""
    获取token
"""
def fetch_token(baidu_struct):
    API_KEY=baidu_struct['API_KEY']
    SECRET_KEY=baidu_struct['SECRET_KEY']
    TOKEN_URL=baidu_struct['TOKEN_URL']

    params = {'grant_type': 'client_credentials',
              'client_id': API_KEY,
              'client_secret': SECRET_KEY}
    post_data = urlencode(params)
    if (IS_PY3):
        post_data = post_data.encode('utf-8')
    req = Request(TOKEN_URL, post_data)
    try:
        f = urlopen(req, timeout=10)
        result_str = f.read()
    except URLError as err:
        print(err)
    if (IS_PY3):
        result_str = result_str.decode()


    result = json.loads(result_str)

    if ('access_token' in result.keys() and 'scope' in result.keys()):
        if not 'brain_all_scope' in result['scope'].split(' '):
            print ('please ensure has check the  ability')
            exit()
        return result['access_token']
    else:
        print ('please overwrite the correct API_KEY and SECRET_KEY')
        exit()

"""
    读取文件
"""
def read_file(image_path):
    f = None
    try:
        f = open(image_path, 'rb')
        return f.read()
    except:
        print('read image file fail')
        return None
    finally:
        if f:
            f.close()


"""
    调用远程服务
"""
def request(url, data):
    req = Request(url, data.encode('utf-8'))
    has_error = False
    try:
        f = urlopen(req)
        result_str = f.read()
        if (IS_PY3):
            result_str = result_str.decode()
        return result_str
    except  URLError as err:
        print(err)


def baidu_ocr(baidu_struct,png_file_path,xls_file_path):
    OCR_URL=baidu_struct['OCR_URL']
    # 获取access token
    token = fetch_token(baidu_struct)

    # 拼接通用文字识别高精度url
    image_url = OCR_URL + "?access_token=" + token

    text = ""


    ################
    # 读取测试图片
    file_content = read_file(png_file_path)

    img_base64 = base64.b64encode(file_content).decode()
    params = {
        "image": img_base64,
        "result_type": "excel",
        "return_excel": "true"
    }
    data = urlencode(params)

    result = request(image_url, data)
    result_json = json.loads(result)
    #print(json.dumps(result_json, indent=2, ensure_ascii=False))
    excel_base64 = result_json['excel_file']
    # 解码为二进制内容
    excel_bytes = base64.b64decode(excel_base64)
    # 保存为xlsx文件
    with open(xls_file_path, 'wb') as f:
        f.write(excel_bytes)

    print("Excel文件已保存为 baidu_table_ocr.xlsx")


def tencent_ocr(tencent_struct,png_file_path,xls_file_path):
    try:
        # 实例化一个认证对象，入参需要传入腾讯云账户 SecretId 和 SecretKey，此处还需注意密钥对的保密
        # 代码泄露可能会导致 SecretId 和 SecretKey 泄露，并威胁账号下所有资源的安全性
        # 以下代码示例仅供参考，建议采用更安全的方式来使用密钥
        # 请参见：https://cloud.tencent.com/document/product/1278/85305
        # 密钥可前往官网控制台 https://console.cloud.tencent.com/cam/capi 进行获取
        # cred = credential.Credential(os.getenv("TENCENTCLOUD_SECRET_ID"), os.getenv("TENCENTCLOUD_SECRET_KEY"))
        # 使用临时密钥示例
        cred = credential.Credential(tencent_struct["SecretId"], tencent_struct["SecretKey"], "Token")
        # 实例化一个http选项，可选的，没有特殊需求可以跳过
        httpProfile = HttpProfile()
        httpProfile.endpoint = "ocr.tencentcloudapi.com"

        # 实例化一个client选项，可选的，没有特殊需求可以跳过
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        # 实例化要请求产品的client对象,clientProfile是可选的
        client = ocr_client.OcrClient(cred, "", clientProfile)

        # 实例化一个请求对象,每个接口都会对应一个request对象
        req = models.RecognizeTableAccurateOCRRequest()

        file_content = read_file(png_file_path)
        img_base64 = base64.b64encode(file_content).decode()

        params = {
            'ImageBase64':img_base64
        }
        req.from_json_string(json.dumps(params))

        # 返回的resp是一个RecognizeTableAccurateOCRResponse的实例，与请求对象对应
        resp = client.RecognizeTableAccurateOCR(req)
        # 输出json格式的字符串回包
        #print(resp.to_json_string())

        excel_base64 = resp.Data
        # 解码为二进制内容
        excel_bytes = base64.b64decode(excel_base64)
        # 保存为xlsx文件
        with open(xls_file_path, 'wb') as f:
            f.write(excel_bytes)

        print("Excel文件已保存为"+xls_file_path+".xlsx")
        return 0
    except TencentCloudSDKException as err:
        print(err)
        return -1


def fix_excel_linebreak(xls_file_path):
    """
    修复Excel文件中的换行问题。

    该函数会将单元格中的换行文本拆分到不同的行中，同时保留原有的合并单元格格式。

    参数:
    xls_file_path (str): Excel文件的路径。
    """
    # 加载Excel工作簿
    wb = openpyxl.load_workbook(xls_file_path)
    # 获取活动工作表
    ws = wb.active

    # 解除已有的单元格合并，以便后续操作
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    # 获取工作表的最大行数和列数
    max_row = ws.max_row
    max_col = ws.max_column

    # # 逆序遍历工作表中的每个单元格
    # for row in range(max_row, 0, -1):
    #     for col in range(1, max_col + 1):
    #         # 获取当前单元格
    #         cell = ws.cell(row=row, column=col)
    #         # 检查单元格值是否包含换行符
    #         if cell.value and isinstance(cell.value, str) and '\n' in cell.value:
    #             # 将单元格值按换行符分割
    #             lines = cell.value.split('\n')
    #             # 当前格写第1行，其余依次写到下方各行
    #             for idx, line in enumerate(lines):
    #                 target_row = row + idx
    #                 # 防止越界，可扩展表
    #                 if target_row > ws.max_row:
    #                     ws.insert_rows(ws.max_row + 1)
    #                 # 将分割后的每行文本写入单元格
    #                 ws.cell(row=target_row, column=col).value = line

    for row in range(max_row, 0, -1):
        # 获取第1列的单元格内容，用于判断是否为表正文
        first_col_value = ws.cell(row=row, column=1).value
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and '\n' in cell.value:
                # 新增：如果本行第1列不是数字，跳过本单元格换行处理
                if not (first_col_value and str(first_col_value).strip().replace('.', '', 1).isdigit()):
                    continue  # 跳过表头/非数字行
                lines = cell.value.split('\n')
                for idx, line in enumerate(lines):
                    target_row = row + idx
                    if target_row > ws.max_row:
                        ws.insert_rows(ws.max_row + 1)
                    ws.cell(row=target_row, column=col).value = line

    # 保存修改后的Excel文件
    wb.save(xls_file_path)


# if __name__ == '__main__':
#     if def_baidu:
#         baidu_struct = {'API_KEY': 'F8Go5rLbjMozTcWE5yClhKCa',
#                         'SECRET_KEY': 'm9Ey40uL6mE4N45Hndr8IdqANImbaVjB',
#                         'OCR_URL': 'https://aip.baidubce.com/rest/2.0/ocr/v1/table',
#                         'TOKEN_URL': 'https://aip.baidubce.com/oauth/2.0/token'}
#         png_file_path='./testpng2.png'
#         xls_file_path= 'baidu_table_ocr.xlsx'
#         baidu_ocr(baidu_struct,png_file_path,xls_file_path)
#
#         print("百度表格识别结束")
#
#     if def_tencent:
#         tencent_struct={'SecretId':'AKIDaKwUq4A5qOiqoa9VfeVBetztdBawHScp',
#                         'SecretKey':'FH92lHbifaXN9pJb7nGj9Uq585y33yro',
#                         }
#         png_file_path='./testpng2.png'
#         xls_file_path= 'tencent_table_ocr.xlsx'
#
#         tencent_ocr(tencent_struct,png_file_path,xls_file_path)
#
#         print("腾讯表格识别结束")
#
#     fix_excel_linebreak(xls_file_path)
#
#     import os


def batch_ocr(ocr_struct, in_dir, out_dir, ocr_func):
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    for fname in os.listdir(in_dir):
        if fname.lower().endswith('.png'):
            png_file_path = os.path.join(in_dir, fname)
            xls_file_path = os.path.join(out_dir, os.path.splitext(fname)[0] + ".xlsx")
            try:
                err=ocr_func(ocr_struct, png_file_path, xls_file_path)
                fix_excel_linebreak(xls_file_path)
                if err==0:
                    print(f"{fname} 转换完成")
                else:
                    #发送出错exception
                    raise Exception(f"OCR处理失败，错误码：{err}，文件：{fname}")
            except Exception as e:
                print(f"{fname} 转换出错: {e}")


### excel combination groups
def is_data_row(row):
    # 判断是否为正文：第一个单元格首字符为数字
    if not row: return False
    v = str(row[0].value).strip()
    return v and v[0].isdigit()

def get_prefix(filename):
    # 返回如A700，A750等前缀
    return filename.split('_')[0]

def find_merge_groups(folder):
    # 返回 {'A700': ['A700_pt1.xlsx', ...], ...}
    files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]
    groups = {}
    for f in files:
        prefix = get_prefix(f)
        groups.setdefault(prefix, []).append(f)
    return groups

def get_header_rows(sheet):
    # 返回表头结束行（正文第一行索引，从1开始）
    for idx, row in enumerate(sheet.iter_rows(min_row=1), 1):
        if is_data_row(row):
            return idx
    return 1

def merge_excels(folder, output_folder):
    groups = find_merge_groups(folder)
    for prefix, files in groups.items():
        # 排序：保证pt1最先，pt2, pt3后续
        files_sorted = sorted(files, key=lambda x: int(x.split('_pt')[1].split('.')[0]))
        wb_out = None
        ws_out = None
        for idx, fname in enumerate(files_sorted):
            path = os.path.join(folder, fname)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            if idx == 0:
                # pt1：全部保留
                wb_out = openpyxl.Workbook()
                ws_out = wb_out.active
                ws_out.title = ws.title
                for row in ws.iter_rows(values_only=False):
                    ws_out.append([cell.value for cell in row])
            else:
                # pt2/pt3...：仅插入正文部分
                header_row = get_header_rows(ws)
                for row in ws.iter_rows(min_row=header_row, values_only=False):
                    ws_out.append([cell.value for cell in row])
        if wb_out:
            out_path = os.path.join(output_folder, f"{prefix}.xlsx")
            wb_out.save(out_path)
            print(f"{prefix} 合并完成，输出: {out_path}")

def combine_large_excel(folder, output_folder_name):

    files = [f for f in os.listdir(folder) if f.endswith('.xls') or f.endswith('.xlsx')]

    # 逐个读入DataFrame，合并（并集模式，自动对齐所有不同表头）
    df_list = []
    for file in files:
        df = pd.read_excel(os.path.join(folder, file))
        df_list.append(df)

    # 并集合并（忽略索引，按所有列对齐，缺失处NaN）
    big_table = pd.concat(df_list, ignore_index=True, sort=False)

    # 保存为新的Excel
    big_table.to_excel(output_folder_name, index=False)
    print('合并完成'+output_folder_name)


if __name__ == '__main__':
    def_convert_png_to_xlsx=False
    def_combine_excels=False
    def_combine_large_excel=True

    in_dir = './datasheets/pngs'
    out_dir = './datasheets/output_excels'
    if def_convert_png_to_xlsx:
        if def_baidu:
            baidu_struct = {'API_KEY': 'F8Go5rLbjMozTcWE5yClhKCa',
                            'SECRET_KEY': 'm9Ey40uL6mE4N45Hndr8IdqANImbaVjB',
                            'OCR_URL': 'https://aip.baidubce.com/rest/2.0/ocr/v1/table',
                            'TOKEN_URL': 'https://aip.baidubce.com/oauth/2.0/token'}
            batch_ocr(baidu_struct, in_dir, out_dir, baidu_ocr)
            print("百度表格批量识别结束")
        if def_tencent:
            tencent_struct={'SecretId':'AKIDaKwUq4A5qOiqoa9VfeVBetztdBawHScp',
                            'SecretKey':'FH92lHbifaXN9pJb7nGj9Uq585y33yro',
                            }
            batch_ocr(tencent_struct, in_dir, out_dir, tencent_ocr)
            print("腾讯表格批量识别结束")

    #combine xls
    combine_xls_dir = './datasheets/combine_xls_unchecked'
    if def_combine_excels:
        merge_excels(out_dir,combine_xls_dir)


    #manual modify
    #combine xls into large sheet
    combine_xls_dir = './datasheets/combine_xls'
    combine_large_excel_path='./datasheets/combine_xls/combined_large_excel/combined_large_excel.xlsx'
    if def_combine_large_excel:
        combine_large_excel(combine_xls_dir,combine_large_excel_path)
    print('end')








